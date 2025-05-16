from dataclasses import dataclass
import datetime
import openpyxl
import time
import requests


@dataclass
class SearchCriteria:
    min_bond_yield: float = 24.05
    max_bond_yield: float = float("inf")
    min_days_to_maturity: float = 0
    max_days_to_maturity: float = float("inf")
    face_units: list[str] | None = ("SUR",)  # Use None if don't care about face unit


class Bond:
    BROKER_FEE = 0.25 / 100

    def __init__(
        self,
        ISIN: str,
        bond_name: str,
        face_value: float,
        coupon_value: float,
        coupon_period: int,
        maturity_date: datetime.date,
        bond_price: float,
        ACI: float,
        face_unit: str,
        is_qualified: bool = False,
    ):
        self.ISIN: str = ISIN
        self.bond_name: str = bond_name
        self.face_value: float = face_value or 0
        self.coupon_value: float = coupon_value or 0
        self.coupon_period: int = coupon_period or float("inf")
        self.maturity_date: datetime.date = maturity_date
        self.bond_price: float = bond_price or float("inf")
        self.ACI: float = ACI
        self.face_unit: str = face_unit
        self.is_qualified: bool = is_qualified

    @classmethod
    def headers(cls):
        return [
            "Наименование",
            "ISIN",
            "Дней до погашения",
            "Доходность к погашению",
            "Валюта",
            "Требуется квалификация",
        ]

    @property
    def as_list(self):
        return [
            self.bond_name,
            self.ISIN,
            self.days_to_maturity,
            self.yield_to_maturity,
            self.face_unit,
            "Да" if self.is_qualified else "Нет",
        ]

    @property
    def days_to_maturity(self):
        return (self.maturity_date - datetime.date.today()).days

    @property
    def yield_to_maturity(self):
        if self.days_to_maturity <= 0:
            return 0

        full_coupons, part_coupon = divmod(self.days_to_maturity, self.coupon_period)

        coupons = full_coupons + bool(part_coupon)
        coupons_income = coupons * self.coupon_value

        clean_price = self.face_value * self.bond_price / 100  # no ACI
        price = clean_price + self.ACI  # current market price
        price *= 1 + self.BROKER_FEE  # including broker fee

        total_income = self.face_value + coupons_income
        rate = (total_income / price - 1) * 100 * 365 / self.days_to_maturity

        return round(rate, 2)


def output_to_excel(bond_list: list[Bond]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(Bond.headers())
    for bond in bond_list:
        ws.append(bond.as_list)
    # Center data
    center_alignment = openpyxl.styles.Alignment(horizontal="center")
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
    ):
        for cell in row:
            cell.alignment = center_alignment
    # Auto-width
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(column_cells[0].col_idx)
        ].width = (length * 1.1)
    # save file
    wb.save(f"{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx")
    return


def LOG(message: str) -> None:
    print(message)


def get_json(url: str) -> dict:
    response = requests.get(url)
    return response.json()


def get_bond_qualification(ISIN: str) -> bool | None:
    url = (
        f"https://iss.moex.com/iss/securities/{ISIN}.json?"
        "iss.meta=off"
        "&iss.only=description&"
        "description.columns=name,title,value&"
    )
    LOG(f"Проверка требования квалификации для {ISIN}.")
    try:
        time.sleep(API_DELAY)
        json_data = get_json(url)
        description_data = json_data["description"]["data"]

        is_qualified_investors_data = next(
            (item for item in description_data if item[0] == "ISQUALIFIEDINVESTORS"),
            None,
        )
        qual_investor_group_data = next(
            (item for item in description_data if item[0] == "QUALINVESTORGROUP"),
            None,
        )

        is_qualified_investors = (
            int(is_qualified_investors_data[2])
            if is_qualified_investors_data and is_qualified_investors_data[2]
            else 0
        )  # По умолчанию 0, если не найдено

        qual_investor_group = (
            qual_investor_group_data[2]
            if qual_investor_group_data and qual_investor_group_data[2]
            else "не определена"
        )  # Текст по умолчанию, если не найден

        if is_qualified_investors == 0:
            LOG(f"Для {ISIN} квалификация для покупки НЕ нужна.")
            return False
        else:
            LOG(
                f"{ISIN} для квалифицированных инвесторов категории: {qual_investor_group}"
            )
            return True
    except Exception as e:
        LOG(f"Ошибка с {ISIN} при получении сведений о необходимой квалификации.")
        return None


def parse_boardgroup(boardgroup: int) -> list[Bond]:
    bonds: list[Bond] = []
    time.sleep(API_DELAY)
    url = (
        f"https://iss.moex.com/iss/engines/stock/markets/bonds/boardgroups/{boardgroup}/securities.json?"
        "iss.dp=comma&iss.meta=off&"
        "iss.only=securities&"
        "securities.columns=ISIN,SHORTNAME,FACEVALUE,COUPONVALUE,COUPONPERIOD,MATDATE,PREVLEGALCLOSEPRICE,ACCRUEDINT,FACEUNIT&"
    )
    LOG(f"Ссылка поиска всех доступных облигаций группы {boardgroup}: {url}.\n\n")

    try:
        json_data: dict = get_json(url)
    except:
        LOG("Ошибка запроса к API.")

    securities_data: list = json_data.get("securities", {}).get("data", [])
    if not (securities_data):
        LOG(
            f"Нет данных c Московской биржи для группы {boardgroup}."
            "Проверьте вручную по ссылке выше.\n\n"
        )
        return []

    securities_data: dict = {item[0]: item for item in securities_data}
    LOG(f"Всего в списке группы {boardgroup} - {len(securities_data)} бумаг.")

    for i, ISIN in enumerate(securities_data, start=1):
        LOG(f"\nСтрока {i} из {len(securities_data)}:")

        # Pasing bond data
        bond_name = str(securities_data[ISIN][1])

        try:
            face_value = float(securities_data[ISIN][2])
        except TypeError:
            LOG("Ошибка при получении номинала облигации.")
            LOG(f"Пропуск {ISIN}")
            continue

        try:
            coupon_value = float(securities_data[ISIN][3])
        except TypeError:
            coupon_value = 0
            LOG("Ошибка при получении номинала купона.")
            LOG(f"Пропуск {ISIN}")
            continue

        try:
            coupon_period = float(securities_data[ISIN][4])
        except TypeError:
            LOG("Ошибка при получении периода купона.")
            LOG(f"Пропуск {ISIN}")
            continue

        try:
            mat_date = securities_data[ISIN][5]
            date_format = "%Y-%m-%d"
            maturity_date = datetime.datetime.strptime(mat_date, date_format).date()
        except ValueError:
            LOG("Ошибка при получении даты погашения облигации.")
            LOG(f"Пропуск {ISIN}.")
            continue

        try:
            bond_price = float(securities_data[ISIN][6])
        except TypeError:
            LOG("Ошибка при получении цены облигации на бирже.")
            LOG(f"Пропуск {ISIN}.")
            continue

        try:
            ACI = float(securities_data[ISIN][7])
        except TypeError:
            LOG("Ошибка при получении накопленного купонного дохода.")
            LOG(f"Пропуск {ISIN}")
            continue

        face_unit = securities_data[ISIN][8]

        bond: Bond = Bond(
            ISIN=ISIN,
            bond_name=bond_name,
            face_value=face_value,
            coupon_value=coupon_value,
            coupon_period=coupon_period,
            maturity_date=maturity_date,
            bond_price=bond_price,
            ACI=ACI,
            face_unit=face_unit,
        )

        # Checking basic search criteria
        condition = (
            search_criteria.min_days_to_maturity
            <= bond.days_to_maturity
            <= search_criteria.max_days_to_maturity
            and search_criteria.min_bond_yield
            <= bond.yield_to_maturity
            <= search_criteria.max_bond_yield
            and (
                search_criteria.face_units
                and (bond.face_unit in search_criteria.face_units)
            )
        )

        if condition:
            LOG(
                f"Условие "
                f"доходности ({search_criteria.min_bond_yield}% <= {bond.yield_to_maturity}% <= {search_criteria.max_bond_yield}%), "
                f"дней до погашения ({search_criteria.min_days_to_maturity} <= {bond.days_to_maturity} <= {search_criteria.max_days_to_maturity}), "
                f"валюта ({bond.face_unit} в {search_criteria.face_units}) "
                f"для {ISIN} прошло."
            )

            # bond.is_qualified = get_bond_qualification(ISIN)

            bonds.append(bond)
        else:
            LOG(f"{bond_name} с ISIN {ISIN} не соответсвует критериям поиска.")
    return bonds


API_DELAY = round(60 / 50, 1)
BOARDGROUPS = [7, 58, 105]
search_criteria = SearchCriteria()

bonds = []
for boardgroup in BOARDGROUPS:
    bonds.extend(parse_boardgroup(boardgroup))

output_to_excel(bonds)