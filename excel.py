import datetime
import logging
import openpyxl

from schemas import *

logger = logging.getLogger("Excel")


class ExcelBook:
    def __init__(self, file_name: str = None, max_save_attempts: int = 5):
        self.file_name = self._normalize_file_name(file_name)
        self.max_save_attempts = max_save_attempts

    def write_bonds(self, bond_list: list[Bond]) -> None:
        """
        Writes given bond_list to excel file.
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(Bond.headers())
        for bond in bond_list:
            ws.append(bond.as_list)

        self._center_worksheet(ws)
        self._auto_width(ws)

        self._save_with_retries(wb)

    @staticmethod
    def _normalize_file_name(file_name: str | None) -> str:
        """
        Normalizes given file name.
        If no file name given - sets default file name to current date `%d.%m.%Y`.
        """
        if file_name is None:
            file_name = datetime.datetime.now().strftime("%d.%m.%Y")
        if not file_name.endswith(".xlsx"):
            file_name += ".xlsx"
        return file_name

    def _center_worksheet(
        self, worksheet: openpyxl.worksheet.worksheet.Worksheet
    ) -> None:
        """
        Centers data in worksheet.
        """
        logger.info("Центрирование ячеек таблицы...")
        center_alignment = openpyxl.styles.Alignment(horizontal="center")
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                cell.alignment = center_alignment

    def _auto_width(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        Sets width to fit content
        """
        logger.info("Подбор ширины ячеек таблицы")
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells if cell.value)
            worksheet.column_dimensions[
                openpyxl.utils.get_column_letter(column_cells[0].col_idx)
            ].width = (length * 1.2)
        return

    def _save_with_retries(self, wb: openpyxl.Workbook) -> None:
        """
        Tries to save file self.max_save_attempts times.
        If file couldn't be saved - raises IOError.
        """
        attempt = 0
        while attempt < self.max_save_attempts:
            try:
                filename = (
                    self.file_name + (f"({attempt})" if attempt else "") + ".xlsx"
                )
                wb.save(filename)
                logger.info(f"Файл сохранен: {filename}.")
                return
            except PermissionError:
                logger.warning(
                    f"Не удалось сохранить файл {filename}, пробую другое имя..."
                )
                attempt += 1
        raise IOError(
            f"Не удалось сохранить файл после {self.max_save_attempts} попыток."
        )
